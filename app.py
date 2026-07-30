import csv
import io
import json
import os
import secrets
import zipfile
from datetime import date
from functools import wraps

import openpyxl
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, abort

import db
import api_compras

UASG_PADRAO = "153010"

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or "chave-dev-local-nao-usar-em-producao"
app.config["MAX_CONTENT_LENGTH"] = 21 * 1024 * 1024  # cobre até 2 anexos de 10MB (pedido + cessão) + folga do multipart

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")

db.init_db()


# ------------------------------------------------------------------- auth --

def requer_login(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("autenticado"):
            return redirect(url_for("login", proximo=request.path))
        return view(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if ADMIN_PASSWORD and secrets.compare_digest(senha, ADMIN_PASSWORD):
            session["autenticado"] = True
            return redirect(request.form.get("proximo") or url_for("dashboard"))
        erro = "Senha incorreta."
    return render_template("login.html", erro=erro, proximo=request.args.get("proximo", ""))


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------- helpers --

def calcular_saldos(conn, uasg=UASG_PADRAO):
    linhas = conn.execute("""
        SELECT
            p.id AS planejamento_id,
            p.numero_ata,
            p.numero_item,
            p.centro_custo,
            p.quantidade_planejada,
            i.descricao,
            i.unidade,
            i.quantidade_homologada,
            a.numero_controle_pncp_compra,
            a.numero_compra AS pregao_numero_compra,
            a.ano_compra AS pregao_ano_compra,
            COALESCE(pc.liberado, FALSE) AS pregao_liberado,
            pc.prazo_final AS pregao_prazo_final,
            COALESCE((
                SELECT SUM(pe.quantidade_solicitada) FROM pedidos pe
                WHERE pe.numero_ata = p.numero_ata AND pe.uasg = p.uasg AND pe.numero_item = p.numero_item
                      AND pe.centro_custo = p.centro_custo AND pe.status IN ('Solicitado', 'Empenhado')
            ), 0) AS usado,
            COALESCE((
                SELECT SUM(pe.quantidade_solicitada) FROM pedidos pe
                WHERE pe.numero_ata = p.numero_ata AND pe.uasg = p.uasg AND pe.numero_item = p.numero_item
                      AND pe.status IN ('Solicitado', 'Empenhado')
            ), 0) AS usado_total_item
        FROM planejamento p
        LEFT JOIN itens i ON i.numero_ata = p.numero_ata AND i.uasg = p.uasg AND i.numero_item = p.numero_item
        LEFT JOIN arps a ON a.numero_ata = p.numero_ata AND a.uasg = p.uasg
        LEFT JOIN pregoes_controle pc ON pc.numero_controle_pncp_compra = a.numero_controle_pncp_compra
                                       AND pc.uasg = p.uasg
        WHERE p.uasg = %s
        ORDER BY p.numero_ata, p.numero_item, p.centro_custo
    """, (uasg,)).fetchall()

    resultado = []
    for linha in linhas:
        d = dict(linha)
        d["saldo"] = d["quantidade_planejada"] - d["usado"]
        d["saldo_item_total"] = (d["quantidade_homologada"] or 0) - d["usado_total_item"]
        resultado.append(d)
    return resultado


def pregao_aberto(linha):
    """Confere se o pregão de uma linha de calcular_saldos() está liberado e dentro do prazo."""
    if not linha.get("pregao_liberado"):
        return False
    prazo = linha.get("pregao_prazo_final")
    if not prazo:
        return True
    return date.today().isoformat() <= str(prazo)


def somar_planejamento_por_item(conn, uasg=UASG_PADRAO):
    linhas = conn.execute("""
        SELECT numero_ata, numero_item, SUM(quantidade_planejada) AS total_planejado
        FROM planejamento WHERE uasg = %s GROUP BY numero_ata, numero_item
    """, (uasg,)).fetchall()
    return {(r["numero_ata"], r["numero_item"]): r["total_planejado"] for r in linhas}


def normaliza_ata(valor):
    return str(valor).strip()


def normaliza_item(valor):
    valor = str(valor).strip()
    if valor.isdigit():
        return valor.zfill(5)
    return valor


def id_campo_quantidade(numero_ata, numero_item):
    return "qtd_" + numero_ata.replace("/", "-") + "_" + numero_item


def campo_planejamento(centro_codigo, numero_ata, numero_item):
    return "pl_" + centro_codigo.replace(" ", "_") + "__" + numero_ata.replace("/", "-") + "_" + numero_item


app.jinja_env.globals["id_campo_quantidade"] = id_campo_quantidade
app.jinja_env.globals["campo_planejamento"] = campo_planejamento


# ------------------------------------------------------------------ rotas --

@app.route("/")
@requer_login
def dashboard():
    conn = db.get_conn()
    try:
        saldos = calcular_saldos(conn)
        centros = [r["codigo"] for r in conn.execute("SELECT codigo FROM centros_custo ORDER BY codigo")]
        filtro_centro = request.args.get("centro_custo", "").strip()
        filtro_busca = request.args.get("busca", "").strip().lower()
        if filtro_centro:
            saldos = [s for s in saldos if s["centro_custo"] == filtro_centro]
        if filtro_busca:
            saldos = [s for s in saldos if filtro_busca in (s["descricao"] or "").lower()
                      or filtro_busca in s["numero_ata"].lower() or filtro_busca in s["numero_item"].lower()]
        total_arps = conn.execute("SELECT COUNT(*) c FROM arps").fetchone()["c"]
        ultima_sync = conn.execute("SELECT MAX(atualizado_em) u FROM itens").fetchone()["u"]
        return render_template("dashboard.html", saldos=saldos, centros=centros,
                                filtro_centro=filtro_centro, filtro_busca=filtro_busca,
                                total_arps=total_arps, ultima_sync=ultima_sync)
    finally:
        conn.close()


@app.route("/sincronizar", methods=["POST"])
@requer_login
def sincronizar():
    uasg = request.form.get("uasg", UASG_PADRAO).strip() or UASG_PADRAO
    try:
        resultado = api_compras.sincronizar(uasg)
        msg = (f"Sincronização concluída: {resultado['arps_sincronizadas']} ARPs, "
               f"{resultado['itens_sincronizados']} itens, "
               f"{resultado['atas_com_empenho_consultadas']} atas verificadas para empenhos.")
        if resultado["avisos"]:
            msg += f" {len(resultado['avisos'])} avisos (ver log do servidor)."
            for a in resultado["avisos"]:
                app.logger.warning(a)
        flash(msg, "sucesso")
    except Exception as e:
        flash(f"Falha na sincronização com a API: {e}", "erro")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/planejamento", methods=["GET"])
@requer_login
def planejamento():
    conn = db.get_conn()
    try:
        itens = conn.execute("""
            SELECT numero_ata, uasg, numero_item, descricao, unidade, quantidade_homologada
            FROM itens WHERE uasg = %s ORDER BY numero_ata, numero_item
        """, (UASG_PADRAO,)).fetchall()
        planos = conn.execute("""
            SELECT * FROM planejamento WHERE uasg = %s ORDER BY numero_ata, numero_item, centro_custo
        """, (UASG_PADRAO,)).fetchall()
        totais = somar_planejamento_por_item(conn, UASG_PADRAO)
        centros = conn.execute("SELECT * FROM centros_custo ORDER BY codigo").fetchall()

        alerta_itens = []
        for it in itens:
            chave = (it["numero_ata"], it["numero_item"])
            total_planejado = totais.get(chave, 0)
            if total_planejado - (it["quantidade_homologada"] or 0) > 1e-6:
                alerta_itens.append({
                    "numero_ata": it["numero_ata"], "numero_item": it["numero_item"],
                    "descricao": it["descricao"], "quantidade_homologada": it["quantidade_homologada"],
                    "total_planejado": total_planejado,
                })

        return render_template("planejamento.html", itens=itens, planos=planos, centros=centros,
                                totais=totais, alerta_itens=alerta_itens)
    finally:
        conn.close()


@app.route("/planejamento/novo", methods=["POST"])
@requer_login
def planejamento_novo():
    numero_ata = normaliza_ata(request.form["numero_ata"])
    numero_item = normaliza_item(request.form["numero_item"])
    centro_custo = request.form["centro_custo"].strip()
    try:
        quantidade = float(request.form["quantidade_planejada"].replace(",", "."))
    except ValueError:
        flash("Quantidade planejada inválida.", "erro")
        return redirect(url_for("planejamento"))

    conn = db.get_conn()
    try:
        item = conn.execute("SELECT 1 FROM itens WHERE numero_ata=%s AND uasg=%s AND numero_item=%s",
                             (numero_ata, UASG_PADRAO, numero_item)).fetchone()
        if not item:
            flash(f"Item {numero_item} da ata {numero_ata} não encontrado. Sincronize a API antes de planejar.", "erro")
            return redirect(url_for("planejamento"))

        ts = db.agora()
        conn.execute("""
            INSERT INTO planejamento (numero_ata, uasg, numero_item, centro_custo, quantidade_planejada, criado_em, atualizado_em)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (numero_ata, uasg, numero_item, centro_custo) DO UPDATE SET
                quantidade_planejada = excluded.quantidade_planejada, atualizado_em = excluded.atualizado_em
        """, (numero_ata, UASG_PADRAO, numero_item, centro_custo, quantidade, ts, ts))
        conn.commit()
        flash(f"Planejamento salvo: ata {numero_ata}, item {numero_item}, {centro_custo} = {quantidade}.", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("planejamento"))


@app.route("/planejamento/<int:plano_id>/excluir", methods=["POST"])
@requer_login
def planejamento_excluir(plano_id):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM planejamento WHERE id=%s", (plano_id,))
        conn.commit()
        flash("Registro de planejamento removido.", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("planejamento"))


@app.route("/centros-custo/novo", methods=["POST"])
@requer_login
def centro_custo_novo():
    codigo = request.form["codigo"].strip()
    nome = request.form["nome"].strip()
    conn = db.get_conn()
    try:
        conn.execute("""
            INSERT INTO centros_custo (codigo, nome, token) VALUES (%s, %s, %s)
            ON CONFLICT (codigo) DO UPDATE SET nome = excluded.nome
        """, (codigo, nome, db.novo_token()))
        conn.commit()
        flash(f"Centro de custo {codigo} salvo.", "sucesso")
    finally:
        conn.close()
    return redirect(request.referrer or url_for("planejamento"))


@app.route("/centros-custo/<codigo>/regenerar-token", methods=["POST"])
@requer_login
def centro_custo_regenerar_token(codigo):
    conn = db.get_conn()
    try:
        conn.execute("UPDATE centros_custo SET token=%s WHERE codigo=%s", (db.novo_token(), codigo))
        conn.commit()
        flash(f"Novo link gerado para {codigo}. O link antigo deixou de funcionar.", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("planejamento"))


# ------------------------------------------------------ planejamento em lote (por pregão) --

@app.route("/planejamento/lote")
@requer_login
def planejamento_lote_lista():
    conn = db.get_conn()
    try:
        pregoes = conn.execute("""
            SELECT a.numero_controle_pncp_compra, a.numero_compra, a.ano_compra, COUNT(*) AS qtd_atas,
                   COALESCE(pc.liberado, FALSE) AS liberado, pc.prazo_final
            FROM arps a
            LEFT JOIN pregoes_controle pc ON pc.numero_controle_pncp_compra = a.numero_controle_pncp_compra
                                           AND pc.uasg = a.uasg
            WHERE a.uasg = %s AND a.numero_controle_pncp_compra IS NOT NULL AND a.numero_controle_pncp_compra <> ''
            GROUP BY a.numero_controle_pncp_compra, a.numero_compra, a.ano_compra, pc.liberado, pc.prazo_final
            ORDER BY a.ano_compra DESC, a.numero_compra DESC
        """, (UASG_PADRAO,)).fetchall()
        return render_template("planejamento_lote_lista.html", pregoes=pregoes)
    finally:
        conn.close()


@app.route("/planejamento/lote/<path:pregao_id>/liberar", methods=["POST"])
@requer_login
def planejamento_lote_liberar(pregao_id):
    liberado = request.form.get("liberado") == "on"
    prazo_final = request.form.get("prazo_final", "").strip() or None
    conn = db.get_conn()
    try:
        conn.execute("""
            INSERT INTO pregoes_controle (numero_controle_pncp_compra, uasg, liberado, prazo_final, atualizado_em)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (numero_controle_pncp_compra, uasg) DO UPDATE SET
                liberado = excluded.liberado, prazo_final = excluded.prazo_final, atualizado_em = excluded.atualizado_em
        """, (pregao_id, UASG_PADRAO, liberado, prazo_final, db.agora()))
        conn.commit()
        if liberado:
            flash(f"Pregão liberado para solicitação{f' até {prazo_final}' if prazo_final else ''}.", "sucesso")
        else:
            flash("Pregão fechado para solicitação.", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("planejamento_lote_lista"))


@app.route("/planejamento/lote/<path:pregao_id>", methods=["GET", "POST"])
@requer_login
def planejamento_lote_grade(pregao_id):
    conn = db.get_conn()
    try:
        pregao = conn.execute("""
            SELECT numero_controle_pncp_compra, numero_compra, ano_compra, COUNT(*) AS qtd_atas
            FROM arps WHERE uasg=%s AND numero_controle_pncp_compra=%s
            GROUP BY numero_controle_pncp_compra, numero_compra, ano_compra
        """, (UASG_PADRAO, pregao_id)).fetchone()
        if not pregao:
            abort(404)

        centros = conn.execute(
            "SELECT * FROM centros_custo ORDER BY ordem NULLS LAST, codigo"
        ).fetchall()

        if request.method == "POST":
            ts = db.agora()
            itens_do_pregao = conn.execute("""
                SELECT i.numero_ata, i.numero_item
                FROM itens i JOIN arps a ON a.numero_ata = i.numero_ata AND a.uasg = i.uasg
                WHERE i.uasg = %s AND a.numero_controle_pncp_compra = %s
            """, (UASG_PADRAO, pregao_id)).fetchall()

            salvos = 0
            for item in itens_do_pregao:
                for centro in centros:
                    campo = campo_planejamento(centro["codigo"], item["numero_ata"], item["numero_item"])
                    valor = request.form.get(campo, "").strip().replace(",", ".")
                    if not valor:
                        continue
                    try:
                        quantidade = float(valor)
                    except ValueError:
                        continue
                    conn.execute("""
                        INSERT INTO planejamento (numero_ata, uasg, numero_item, centro_custo, quantidade_planejada, criado_em, atualizado_em)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (numero_ata, uasg, numero_item, centro_custo) DO UPDATE SET
                            quantidade_planejada = excluded.quantidade_planejada, atualizado_em = excluded.atualizado_em
                    """, (item["numero_ata"], UASG_PADRAO, item["numero_item"], centro["codigo"], quantidade, ts, ts))
                    salvos += 1
            conn.commit()
            flash(f"{salvos} célula(s) salva(s).", "sucesso")
            return redirect(url_for("planejamento_lote_grade", pregao_id=pregao_id))

        itens = conn.execute("""
            SELECT i.numero_ata, i.numero_item, i.descricao, i.unidade AS fornecedor, i.quantidade_homologada
            FROM itens i JOIN arps a ON a.numero_ata = i.numero_ata AND a.uasg = i.uasg
            WHERE i.uasg = %s AND a.numero_controle_pncp_compra = %s
            ORDER BY i.numero_ata, i.numero_item
        """, (UASG_PADRAO, pregao_id)).fetchall()

        planos_existentes = conn.execute("""
            SELECT p.numero_ata, p.numero_item, p.centro_custo, p.quantidade_planejada
            FROM planejamento p JOIN arps a ON a.numero_ata = p.numero_ata AND a.uasg = p.uasg
            WHERE p.uasg = %s AND a.numero_controle_pncp_compra = %s
        """, (UASG_PADRAO, pregao_id)).fetchall()
        valores = {(p["numero_ata"], p["numero_item"], p["centro_custo"]): p["quantidade_planejada"]
                   for p in planos_existentes}

        return render_template("planejamento_lote_grade.html", pregao=pregao, centros=centros,
                                itens=itens, valores=valores)
    finally:
        conn.close()


# ------------------------------------------------------------ novo pedido --
# Ferramenta interna da DIARP para lançamento em lote via planilha (sem PDF).
# O fluxo principal para os centros de custo é /solicitar/<token>, mais abaixo.

COLUNAS_ESPERADAS = ["centro_custo", "numero_ata", "numero_item", "quantidade"]


def ler_planilha_pedido(file_storage):
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = wb.active
    cabecalho = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    indices = {}
    for col in COLUNAS_ESPERADAS:
        if col not in cabecalho:
            raise ValueError(f"Coluna obrigatória '{col}' não encontrada na planilha. "
                              f"Cabeçalho esperado: {', '.join(COLUNAS_ESPERADAS)}.")
        indices[col] = cabecalho.index(col)

    linhas = []
    for row in ws.iter_rows(min_row=2):
        valores = [c.value for c in row]
        if all(v is None or str(v).strip() == "" for v in valores):
            continue
        centro_custo = str(valores[indices["centro_custo"]] or "").strip()
        numero_ata = normaliza_ata(valores[indices["numero_ata"]])
        numero_item = normaliza_item(valores[indices["numero_item"]])
        try:
            quantidade = float(str(valores[indices["quantidade"]]).replace(",", "."))
        except (TypeError, ValueError):
            quantidade = None
        linhas.append({
            "centro_custo": centro_custo, "numero_ata": numero_ata,
            "numero_item": numero_item, "quantidade": quantidade,
        })
    return linhas


@app.route("/novo-pedido", methods=["GET"])
@requer_login
def novo_pedido_form():
    return render_template("novo_pedido.html", preview=None)


@app.route("/novo-pedido/analisar", methods=["POST"])
@requer_login
def novo_pedido_analisar():
    arquivo = request.files.get("arquivo")
    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo .xlsx para enviar.", "erro")
        return redirect(url_for("novo_pedido_form"))

    try:
        linhas = ler_planilha_pedido(arquivo)
    except Exception as e:
        flash(f"Não foi possível ler a planilha: {e}", "erro")
        return redirect(url_for("novo_pedido_form"))

    if not linhas:
        flash("A planilha não contém linhas de pedido.", "erro")
        return redirect(url_for("novo_pedido_form"))

    conn = db.get_conn()
    try:
        saldos = {(s["numero_ata"], s["numero_item"], s["centro_custo"]): s for s in calcular_saldos(conn)}
        preview = []
        for linha in linhas:
            chave = (linha["numero_ata"], linha["numero_item"], linha["centro_custo"])
            saldo_atual = saldos.get(chave)
            item_info = conn.execute(
                "SELECT descricao, unidade FROM itens WHERE numero_ata=%s AND uasg=%s AND numero_item=%s",
                (linha["numero_ata"], UASG_PADRAO, linha["numero_item"])).fetchone()

            p = dict(linha)
            p["descricao"] = item_info["descricao"] if item_info else None
            p["saldo_disponivel"] = None
            p["saldo_resultante"] = None
            if linha["quantidade"] is None:
                p["situacao"] = "erro"
                p["mensagem"] = "Quantidade inválida ou vazia."
            elif not item_info:
                p["situacao"] = "erro"
                p["mensagem"] = "Item não encontrado na ata (sincronize a API)."
            elif saldo_atual is None:
                p["situacao"] = "erro"
                p["mensagem"] = "Sem planejamento cadastrado para este item/centro de custo."
            else:
                p["saldo_disponivel"] = saldo_atual["saldo"]
                p["saldo_resultante"] = saldo_atual["saldo"] - linha["quantidade"]
                if p["saldo_resultante"] < -1e-6:
                    p["situacao"] = "estourado"
                    p["mensagem"] = f"Excede o saldo disponível em {abs(p['saldo_resultante']):.2f}."
                else:
                    p["situacao"] = "ok"
                    p["mensagem"] = "Dentro do saldo planejado."
            preview.append(p)

        tem_erro_bloqueante = any(p["situacao"] == "erro" for p in preview)
        payload = json.dumps(preview, ensure_ascii=False)
        return render_template("novo_pedido.html", preview=preview, payload=payload,
                                arquivo_origem=arquivo.filename, tem_erro_bloqueante=tem_erro_bloqueante)
    finally:
        conn.close()


@app.route("/novo-pedido/confirmar", methods=["POST"])
@requer_login
def novo_pedido_confirmar():
    payload = request.form.get("payload")
    arquivo_origem = request.form.get("arquivo_origem", "")
    observacao = request.form.get("observacao", "").strip()
    try:
        linhas = json.loads(payload)
    except (TypeError, json.JSONDecodeError):
        flash("Sessão de upload expirada, envie a planilha novamente.", "erro")
        return redirect(url_for("novo_pedido_form"))

    linhas_validas = [p for p in linhas if p.get("situacao") != "erro"]
    if not linhas_validas:
        flash("Nenhuma linha válida para registrar.", "erro")
        return redirect(url_for("novo_pedido_form"))

    conn = db.get_conn()
    try:
        ts = db.agora()
        por_centro = {}
        for p in linhas_validas:
            por_centro.setdefault(p["centro_custo"], []).append(p)

        gravados = 0
        for centro_custo, linhas_centro in por_centro.items():
            solicitacao = conn.execute("""
                INSERT INTO solicitacoes (centro_custo, data_solicitacao, nome_arquivo_pdf, conteudo_pdf, observacao, criado_em)
                VALUES (%s, %s, %s, NULL, %s, %s) RETURNING id
            """, (centro_custo, date.today().isoformat(), arquivo_origem, observacao, ts)).fetchone()
            solicitacao_id = solicitacao["id"]

            for p in linhas_centro:
                conn.execute("""
                    INSERT INTO pedidos (solicitacao_id, data_pedido, numero_ata, uasg, numero_item, centro_custo,
                                          quantidade_solicitada, status, criado_em, atualizado_em)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'Solicitado', %s, %s)
                """, (solicitacao_id, date.today().isoformat(), p["numero_ata"], UASG_PADRAO, p["numero_item"],
                      centro_custo, p["quantidade"], ts, ts))
                gravados += 1
        conn.commit()
        flash(f"{gravados} pedido(s) registrado(s) com status 'Solicitado'.", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("pedidos"))


# ---------------------------------------------------------- solicitar (link do centro de custo) --

MAX_TAMANHO_ARQUIVO = 10 * 1024 * 1024  # 10MB por arquivo (não por request — ver MAX_CONTENT_LENGTH acima)


@app.route("/solicitar/<token>", methods=["GET"])
def solicitar_lista(token):
    conn = db.get_conn()
    try:
        centro = conn.execute("SELECT codigo, nome FROM centros_custo WHERE token=%s", (token,)).fetchone()
        if not centro:
            abort(404)
        itens = [s for s in calcular_saldos(conn) if s["centro_custo"] == centro["codigo"] and pregao_aberto(s)]

        pregoes_map = {}
        for it in itens:
            pid = it["numero_controle_pncp_compra"]
            if not pid:
                continue
            info = pregoes_map.setdefault(pid, {
                "pregao_id": pid,
                "numero_compra": it["pregao_numero_compra"],
                "ano_compra": it["pregao_ano_compra"],
                "prazo_final": it["pregao_prazo_final"],
                "qtd_itens": 0,
            })
            info["qtd_itens"] += 1
        pregoes = sorted(pregoes_map.values(), key=lambda p: p["prazo_final"] or "9999-99-99")

        return render_template("solicitar_lista.html", centro=centro, pregoes=pregoes, token=token)
    finally:
        conn.close()


@app.route("/solicitar/<token>/<path:pregao_id>", methods=["GET"])
def solicitar_form(token, pregao_id):
    conn = db.get_conn()
    try:
        centro = conn.execute("SELECT codigo, nome FROM centros_custo WHERE token=%s", (token,)).fetchone()
        if not centro:
            abort(404)
        itens = [s for s in calcular_saldos(conn)
                 if s["centro_custo"] == centro["codigo"] and pregao_aberto(s)
                    and s["numero_controle_pncp_compra"] == pregao_id]
        if not itens:
            abort(404)
        pregao_label = f'{itens[0]["pregao_numero_compra"]}/{itens[0]["pregao_ano_compra"]}'
        return render_template("solicitar.html", centro=centro, itens=itens, token=token,
                                pregao_id=pregao_id, pregao_label=pregao_label)
    finally:
        conn.close()


@app.route("/solicitar/<token>/<path:pregao_id>", methods=["POST"])
def solicitar_confirmar(token, pregao_id):
    conn = db.get_conn()
    try:
        centro = conn.execute("SELECT codigo, nome FROM centros_custo WHERE token=%s", (token,)).fetchone()
        if not centro:
            abort(404)

        pdf = request.files.get("pdf")
        if not pdf or pdf.filename == "":
            flash("Anexe o pedido assinado em PDF antes de enviar.", "erro")
            return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))
        if not pdf.filename.lower().endswith(".pdf"):
            flash("O anexo do pedido precisa ser um arquivo .pdf.", "erro")
            return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))
        conteudo_pdf = pdf.read()
        if len(conteudo_pdf) > MAX_TAMANHO_ARQUIVO:
            flash(f"O PDF do pedido tem {len(conteudo_pdf) / 1024 / 1024:.1f} MB — o limite é 10 MB por arquivo.", "erro")
            return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))

        cessao = request.files.get("cessao")
        nome_cessao, conteudo_cessao = None, None
        if cessao and cessao.filename:
            if not cessao.filename.lower().endswith(".pdf"):
                flash("O anexo de cessão precisa ser um arquivo .pdf.", "erro")
                return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))
            conteudo_cessao = cessao.read()
            if len(conteudo_cessao) > MAX_TAMANHO_ARQUIVO:
                flash(f"O PDF de cessão tem {len(conteudo_cessao) / 1024 / 1024:.1f} MB — o limite é 10 MB por arquivo.", "erro")
                return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))
            nome_cessao = cessao.filename

        saldos_centro = [s for s in calcular_saldos(conn)
                          if s["centro_custo"] == centro["codigo"] and pregao_aberto(s)
                             and s["numero_controle_pncp_compra"] == pregao_id]

        # Reenvio: o que já está "Solicitado" deste centro pra este item vai ser substituído
        # (não somado), então devolve essa quantidade — tanto ao saldo do centro quanto ao saldo
        # total do item na ata — antes de validar o novo valor.
        pendentes_centro = conn.execute("""
            SELECT numero_ata, numero_item, SUM(quantidade_solicitada) AS total
            FROM pedidos WHERE uasg=%s AND centro_custo=%s AND status='Solicitado'
            GROUP BY numero_ata, numero_item
        """, (UASG_PADRAO, centro["codigo"])).fetchall()
        pendentes_centro_map = {(p["numero_ata"], p["numero_item"]): p["total"] for p in pendentes_centro}

        pedidos_a_criar = []
        erros = []
        precisa_cessao = False
        for saldo in saldos_centro:
            campo = id_campo_quantidade(saldo["numero_ata"], saldo["numero_item"])
            valor = request.form.get(campo, "").strip().replace(",", ".")
            if not valor:
                continue
            try:
                quantidade = float(valor)
            except ValueError:
                erros.append(f"Quantidade inválida para o item {saldo['numero_item']} (ata {saldo['numero_ata']}).")
                continue
            if quantidade <= 0:
                continue

            pendente_proprio = pendentes_centro_map.get((saldo["numero_ata"], saldo["numero_item"]), 0)
            saldo_centro_efetivo = saldo["saldo"] + pendente_proprio
            saldo_item_efetivo = saldo["saldo_item_total"] + pendente_proprio

            if quantidade - saldo_item_efetivo > 1e-6:
                erros.append(
                    f"Item {saldo['numero_item']} (ata {saldo['numero_ata']}): pedido de {quantidade:.2f} "
                    f"excede o saldo total do item na ata ({saldo_item_efetivo:.2f}), mesmo com cessão."
                )
                continue
            if quantidade - saldo_centro_efetivo > 1e-6:
                precisa_cessao = True

            pedidos_a_criar.append((saldo["numero_ata"], saldo["numero_item"], quantidade))

        if erros:
            for e in erros:
                flash(e, "erro")
            return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))

        if not pedidos_a_criar:
            flash("Informe ao menos uma quantidade para enviar o pedido.", "erro")
            return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))

        if precisa_cessao and not conteudo_cessao:
            flash("Pelo menos um item está acima do que este centro de custo planejou — "
                  "anexe o documento de cessão para enviar.", "erro")
            return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))

        ts = db.agora()
        solicitacao = conn.execute("""
            INSERT INTO solicitacoes (centro_custo, data_solicitacao, nome_arquivo_pdf, conteudo_pdf,
                                       nome_arquivo_cessao, conteudo_cessao_pdf,
                                       numero_controle_pncp_compra, observacao, criado_em)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (centro["codigo"], date.today().isoformat(), pdf.filename, conteudo_pdf,
              nome_cessao, conteudo_cessao, pregao_id,
              request.form.get("observacao", "").strip(), ts)).fetchone()
        solicitacao_id = solicitacao["id"]

        for numero_ata, numero_item, quantidade in pedidos_a_criar:
            # Reenvio antes do prazo: o pedido anterior ainda não decidido (Solicitado) é substituído
            # pelo novo, sem apagar o histórico nem o PDF antigo. Itens já Empenhado/Não Empenhado
            # (decisão já tomada pela DIARP) não são tocados.
            conn.execute("""
                UPDATE pedidos SET status='Substituído', atualizado_em=%s
                WHERE numero_ata=%s AND uasg=%s AND numero_item=%s AND centro_custo=%s AND status='Solicitado'
            """, (ts, numero_ata, UASG_PADRAO, numero_item, centro["codigo"]))
            conn.execute("""
                INSERT INTO pedidos (solicitacao_id, data_pedido, numero_ata, uasg, numero_item, centro_custo,
                                      quantidade_solicitada, status, criado_em, atualizado_em)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 'Solicitado', %s, %s)
            """, (solicitacao_id, date.today().isoformat(), numero_ata, UASG_PADRAO, numero_item,
                  centro["codigo"], quantidade, ts, ts))
        conn.commit()

        msg = f"Pedido enviado com sucesso: {len(pedidos_a_criar)} item(ns) registrado(s)."
        if precisa_cessao:
            msg += " Documento de cessão anexado."
        flash(msg, "sucesso")
        return redirect(url_for("solicitar_form", token=token, pregao_id=pregao_id))
    finally:
        conn.close()


# ---------------------------------------------------------------- pedidos --

@app.route("/pedidos")
@requer_login
def pedidos():
    conn = db.get_conn()
    try:
        filtro_status = request.args.get("status", "").strip()
        query = """
            SELECT pe.*, i.descricao,
                   s.nome_arquivo_pdf, (s.conteudo_pdf IS NOT NULL) AS tem_pdf,
                   s.nome_arquivo_cessao, (s.conteudo_cessao_pdf IS NOT NULL) AS tem_cessao
            FROM pedidos pe
            LEFT JOIN itens i ON i.numero_ata = pe.numero_ata AND i.uasg = pe.uasg AND i.numero_item = pe.numero_item
            LEFT JOIN solicitacoes s ON s.id = pe.solicitacao_id
            WHERE pe.uasg = %s
        """
        params = [UASG_PADRAO]
        if filtro_status:
            query += " AND pe.status = %s"
            params.append(filtro_status)
        query += " ORDER BY pe.data_pedido DESC, pe.id DESC"
        lista = conn.execute(query, params).fetchall()
        return render_template("pedidos.html", pedidos=lista, status_validos=db.STATUS_VALIDOS,
                                filtro_status=filtro_status)
    finally:
        conn.close()


@app.route("/pedidos/<int:pedido_id>/status", methods=["POST"])
@requer_login
def pedido_status(pedido_id):
    novo_status = request.form.get("status")
    if novo_status not in db.STATUS_VALIDOS:
        flash("Status inválido.", "erro")
        return redirect(url_for("pedidos"))
    conn = db.get_conn()
    try:
        conn.execute("UPDATE pedidos SET status=%s, atualizado_em=%s WHERE id=%s",
                     (novo_status, db.agora(), pedido_id))
        conn.commit()
        flash(f"Pedido #{pedido_id} atualizado para '{novo_status}'.", "sucesso")
    finally:
        conn.close()
    return redirect(request.referrer or url_for("pedidos"))


@app.route("/solicitacoes/<int:solicitacao_id>/pdf")
@requer_login
def solicitacao_pdf(solicitacao_id):
    conn = db.get_conn()
    try:
        row = conn.execute("SELECT nome_arquivo_pdf, conteudo_pdf FROM solicitacoes WHERE id=%s",
                            (solicitacao_id,)).fetchone()
        if not row or not row["conteudo_pdf"]:
            abort(404)
        conteudo = bytes(row["conteudo_pdf"])
        return send_file(io.BytesIO(conteudo), as_attachment=True,
                          download_name=row["nome_arquivo_pdf"] or f"solicitacao_{solicitacao_id}.pdf",
                          mimetype="application/pdf")
    finally:
        conn.close()


@app.route("/solicitacoes/<int:solicitacao_id>/cessao")
@requer_login
def solicitacao_cessao_pdf(solicitacao_id):
    conn = db.get_conn()
    try:
        row = conn.execute("SELECT nome_arquivo_cessao, conteudo_cessao_pdf FROM solicitacoes WHERE id=%s",
                            (solicitacao_id,)).fetchone()
        if not row or not row["conteudo_cessao_pdf"]:
            abort(404)
        conteudo = bytes(row["conteudo_cessao_pdf"])
        return send_file(io.BytesIO(conteudo), as_attachment=True,
                          download_name=row["nome_arquivo_cessao"] or f"cessao_{solicitacao_id}.pdf",
                          mimetype="application/pdf")
    finally:
        conn.close()


# ------------------------------------------------------------- conferência --

@app.route("/conferencia")
@requer_login
def conferencia():
    conn = db.get_conn()
    try:
        linhas = conn.execute("""
            SELECT
                i.numero_ata, i.numero_item, i.descricao, i.quantidade_homologada,
                COALESCE(e.quantidade_empenhada, 0) AS empenhado_api,
                COALESCE((
                    SELECT SUM(pe.quantidade_solicitada) FROM pedidos pe
                    WHERE pe.numero_ata = i.numero_ata AND pe.uasg = i.uasg AND pe.numero_item = i.numero_item
                          AND pe.status = 'Empenhado'
                ), 0) AS empenhado_controle,
                e.atualizado_em AS empenho_atualizado_em
            FROM itens i
            LEFT JOIN empenhos e ON e.numero_ata = i.numero_ata AND e.uasg = i.uasg AND e.numero_item = i.numero_item
            WHERE i.uasg = %s
              AND EXISTS (
                  SELECT 1 FROM pedidos pe
                  WHERE pe.numero_ata = i.numero_ata AND pe.uasg = i.uasg AND pe.numero_item = i.numero_item
              )
            ORDER BY i.numero_ata, i.numero_item
        """, (UASG_PADRAO,)).fetchall()

        divergentes = []
        conferidos = []
        for linha in linhas:
            d = dict(linha)
            d["diferenca"] = d["empenhado_controle"] - d["empenhado_api"]
            if abs(d["diferenca"]) > 1e-6:
                divergentes.append(d)
            elif d["empenhado_api"] > 0 or d["empenhado_controle"] > 0:
                conferidos.append(d)

        return render_template("conferencia.html", divergentes=divergentes, conferidos=conferidos)
    finally:
        conn.close()


@app.route("/novo-pedido/modelo.xlsx")
@requer_login
def modelo_pedido_xlsx():
    conn = db.get_conn()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedido"
        ws.append(COLUNAS_ESPERADAS)
        wb.create_sheet("Itens_disponiveis")
        ws2 = wb["Itens_disponiveis"]
        ws2.append(["numero_ata", "numero_item", "descricao", "quantidade_homologada"])
        for it in conn.execute("SELECT numero_ata, numero_item, descricao, quantidade_homologada FROM itens WHERE uasg=%s ORDER BY numero_ata, numero_item", (UASG_PADRAO,)):
            ws2.append([it["numero_ata"], it["numero_item"], it["descricao"], it["quantidade_homologada"]])
        wb.create_sheet("Centros_custo")
        ws3 = wb["Centros_custo"]
        ws3.append(["codigo", "nome"])
        for c in conn.execute("SELECT codigo, nome FROM centros_custo ORDER BY codigo"):
            ws3.append([c["codigo"], c["nome"]])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name="modelo_pedido.xlsx",
                          mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        conn.close()


# -------------------------------------------------------------- manutenção --
# Expurgo de PDF vencido, exportação/exclusão de atas com vigência encerrada e backup de segurança.
# Tudo disparado manualmente por botão — sem depender de agendador externo.

SQL_PDFS_ELEGIVEIS = """
    SELECT s.id, s.centro_custo, s.nome_arquivo_pdf, s.data_solicitacao, MAX(pc.prazo_final) AS prazo_mais_recente
    FROM solicitacoes s
    JOIN pedidos pe ON pe.solicitacao_id = s.id
    JOIN arps a ON a.numero_ata = pe.numero_ata AND a.uasg = pe.uasg
    LEFT JOIN pregoes_controle pc ON pc.numero_controle_pncp_compra = a.numero_controle_pncp_compra AND pc.uasg = a.uasg
    WHERE s.conteudo_pdf IS NOT NULL OR s.conteudo_cessao_pdf IS NOT NULL
    GROUP BY s.id, s.centro_custo, s.nome_arquivo_pdf, s.data_solicitacao
    HAVING MAX(pc.prazo_final) IS NOT NULL AND MAX(pc.prazo_final)::date <= (CURRENT_DATE - INTERVAL '1 month')
"""


def _linhas_para_csv_bytes(linhas):
    buf = io.StringIO()
    if linhas:
        writer = csv.DictWriter(buf, fieldnames=list(linhas[0].keys()))
        writer.writeheader()
        for linha in linhas:
            writer.writerow(dict(linha))
    return buf.getvalue().encode("utf-8-sig")


@app.route("/manutencao")
@requer_login
def manutencao():
    conn = db.get_conn()
    try:
        pdfs_elegiveis = conn.execute(SQL_PDFS_ELEGIVEIS).fetchall()

        atas_vencidas = conn.execute("""
            SELECT numero_ata, data_vigencia_fim,
                   (data_vigencia_fim::date <= CURRENT_DATE - INTERVAL '1 month') AS elegivel_exclusao
            FROM arps
            WHERE uasg = %s AND data_vigencia_fim IS NOT NULL AND data_vigencia_fim <> ''
                  AND data_vigencia_fim::date <= CURRENT_DATE
            ORDER BY data_vigencia_fim
        """, (UASG_PADRAO,)).fetchall()

        return render_template("manutencao.html", pdfs_elegiveis=pdfs_elegiveis, atas_vencidas=atas_vencidas)
    finally:
        conn.close()


@app.route("/manutencao/pdfs/expurgar", methods=["POST"])
@requer_login
def manutencao_expurgar_pdfs():
    conn = db.get_conn()
    try:
        ids = [r["id"] for r in conn.execute(SQL_PDFS_ELEGIVEIS).fetchall()]
        if ids:
            conn.execute("UPDATE solicitacoes SET conteudo_pdf = NULL, conteudo_cessao_pdf = NULL WHERE id = ANY(%s)", (ids,))
            conn.commit()
        flash(f"{len(ids)} PDF(s) expurgado(s) (mais de 1 mês após o prazo do pregão).", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("manutencao"))


@app.route("/manutencao/atas/<path:numero_ata>/csv")
@requer_login
def manutencao_ata_csv(numero_ata):
    conn = db.get_conn()
    try:
        linhas = conn.execute("""
            SELECT i.numero_ata, i.numero_item, i.descricao, p.centro_custo,
                   COALESCE(p.quantidade_planejada, 0) AS quantidade_planejada,
                   COALESCE((
                       SELECT SUM(pe.quantidade_solicitada) FROM pedidos pe
                       WHERE pe.numero_ata = i.numero_ata AND pe.uasg = i.uasg AND pe.numero_item = i.numero_item
                             AND pe.centro_custo = p.centro_custo AND pe.status = 'Empenhado'
                   ), 0) AS empenhado,
                   COALESCE((
                       SELECT SUM(pe.quantidade_solicitada) FROM pedidos pe
                       WHERE pe.numero_ata = i.numero_ata AND pe.uasg = i.uasg AND pe.numero_item = i.numero_item
                             AND pe.centro_custo = p.centro_custo AND pe.status = 'Solicitado'
                   ), 0) AS solicitado_pendente
            FROM itens i
            LEFT JOIN planejamento p ON p.numero_ata = i.numero_ata AND p.uasg = i.uasg AND p.numero_item = i.numero_item
            WHERE i.numero_ata = %s AND i.uasg = %s
            ORDER BY i.numero_item, p.centro_custo
        """, (numero_ata, UASG_PADRAO)).fetchall()

        conteudo = _linhas_para_csv_bytes(linhas)
        nome = f"ata_{numero_ata.replace('/', '-')}.csv"
        return send_file(io.BytesIO(conteudo), as_attachment=True, download_name=nome, mimetype="text/csv")
    finally:
        conn.close()


@app.route("/manutencao/atas/<path:numero_ata>/excluir", methods=["POST"])
@requer_login
def manutencao_ata_excluir(numero_ata):
    conn = db.get_conn()
    try:
        elegivel = conn.execute("""
            SELECT (data_vigencia_fim::date <= CURRENT_DATE - INTERVAL '1 month') AS ok
            FROM arps WHERE numero_ata=%s AND uasg=%s AND data_vigencia_fim IS NOT NULL AND data_vigencia_fim <> ''
        """, (numero_ata, UASG_PADRAO)).fetchone()
        if not elegivel or not elegivel["ok"]:
            flash("Esta ata ainda não passou 1 mês do fim da vigência — exclusão bloqueada.", "erro")
            return redirect(url_for("manutencao"))

        conn.execute("DELETE FROM pedidos WHERE numero_ata=%s AND uasg=%s", (numero_ata, UASG_PADRAO))
        conn.execute("DELETE FROM planejamento WHERE numero_ata=%s AND uasg=%s", (numero_ata, UASG_PADRAO))
        conn.execute("DELETE FROM empenhos WHERE numero_ata=%s AND uasg=%s", (numero_ata, UASG_PADRAO))
        conn.execute("DELETE FROM itens WHERE numero_ata=%s AND uasg=%s", (numero_ata, UASG_PADRAO))
        conn.execute("DELETE FROM arps WHERE numero_ata=%s AND uasg=%s", (numero_ata, UASG_PADRAO))
        conn.commit()
        flash(f"Dados da ata {numero_ata} excluídos do sistema.", "sucesso")
    finally:
        conn.close()
    return redirect(url_for("manutencao"))


@app.route("/manutencao/backup.zip")
@requer_login
def manutencao_backup():
    conn = db.get_conn()
    try:
        planejamento_rows = conn.execute(
            "SELECT * FROM planejamento ORDER BY numero_ata, numero_item, centro_custo").fetchall()
        pedidos_rows = conn.execute("SELECT * FROM pedidos ORDER BY id").fetchall()

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("planejamento.csv", _linhas_para_csv_bytes(planejamento_rows))
            zf.writestr("pedidos.csv", _linhas_para_csv_bytes(pedidos_rows))
        buf.seek(0)
        nome = f"backup_{date.today().isoformat()}.zip"
        return send_file(buf, as_attachment=True, download_name=nome, mimetype="application/zip")
    finally:
        conn.close()


if __name__ == "__main__":
    # use_reloader=False: o watcher padrão fica reiniciando o processo sozinho neste ambiente
    # (detecta mudanças espúrias em arquivos do site-packages), derrubando requisições em andamento.
    app.run(debug=True, use_reloader=False, port=5057)
